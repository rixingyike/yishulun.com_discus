from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import time
import openpyxl
from openpyxl.styles import Font
import os
import sys

# 导出独立文档：pyinstaller --onefile req5.py

# ChromeDriver路径
# 判断是否为打包后的环境
if hasattr(sys, '_MEIPASS'):
    driver_path = os.path.join(sys._MEIPASS, 'chromedriver.exe')
else:
    driver_path = 'chromedriver.exe'

# 初始化WebDriver
options = webdriver.ChromeOptions()
options.add_argument('--headless')  # 无头模式
options.add_argument('--disable-gpu')
options.add_argument('--no-sandbox')

service = Service(driver_path)
driver = webdriver.Chrome(service=service, options=options)

url = 'https://note.mowen.cn/note/detail?noteUuid=j_b4yxlm1QSdOCrZ3d8zH'

try:
    # 打开网页
    driver.get(url)
    
    # 等待页面加载
    time.sleep(3)  # 根据需要调整等待时间

    # 循环滚动到页面底部，直到滚动条无法继续滚动为止
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    # while True:
    #     # 使用JavaScript将页面滚动到底部
    #     driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    #     # 等待一段时间，让页面加载完成
    #     time.sleep(.5)
    #     # 检查是否已经滚动到底部
    #     if driver.execute_script("return window.pageYOffset + window.innerHeight >= document.body.scrollHeight"):
    #         break

    # 查找所有的评论
    comments = driver.find_elements(By.CSS_SELECTOR, 'div.comment-item')
    # print(comments)

    # 存储结果的列表
    results = []

    # 创建一个新的工作簿和工作表
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    # 添加表头
    ws.append(["用户", "笔记", "点赞数"])

    for comment in comments:
        # 提取用户名
        username = comment.find_element(By.CSS_SELECTOR, 'div.name').text.strip()
        # print(username)
        
        # 提取点赞数
        like_count = comment.find_element(By.CSS_SELECTOR, 'view.vote-count').text.strip()
        
        # 查找评论中的所有链接
        note_links = comment.find_elements(By.CSS_SELECTOR, 'div.comment-note-container')
        # print(len(note_links))

        if len(note_links) > 0:
            for link in note_links:
              note_uuid = link.get_attribute('data-uuid')
              link_name = link.find_element(By.CSS_SELECTOR, 'div.comment-note-title').text.strip()
              href = 'https://note.mowen.cn/note/detail?noteUuid=' + note_uuid
              # print(note_uuid, link_name)
              
              # 只保存包括链接的评论
              results.append((username, link_name, href, like_count, note_uuid))

              # 插入数据
              ws.append([username, link_name, like_count])

              # 设置超链接
              cell = ws.cell(row=ws.max_row, column=2)
              cell.hyperlink = href
              cell.font = Font(color="0000FF", underline="single")
        
    
    # 输出结果
    for result in results:
      # 切片限制result[1]和result[2]的最大长度为10个字符
      text1 = result[0][:10] + ('...' if len(result[0]) > 10 else '')
      text2 = result[1][:10] + ('...' if len(result[1]) > 10 else '')
      # 使用制表符来对齐输出文本
      print(f"{text1}\t{text2}\t{result[3]}\t{result[2]}")

    print("--------------------------------")
    print("count：{}".format(len(results)))

    # 保存工作簿
    wb.save("results.xlsx")

finally:
    driver.quit()
